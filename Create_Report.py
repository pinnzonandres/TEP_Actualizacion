'''
- Script Create_Report.py
- Script encargado de crear el reporte financiero de excel con los diferentes agregados y diseños
- Autor: Wilson Andrés Pinzón (wilson.pinzon@prosperidadsocial.gov.co)
- Fecha_Actualización: 23/08/22
'''

# Importamos las librerias necesarias
import pandas as pd
from openpyxl.styles import PatternFill #Función agregada de la libreria openpyxl para poder añadir rellenos de color en una celda de excel
from datetime import datetime
import os

# Cargamos un diccionario con el equivalente del número del mes según su nombre
month_dict = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 
              4: 'Abril', 5: 'Mayo', 6: 'Junio', 
              7: 'Julio', 8: 'Agosto', 9: 'Septiembre',
              10: 'Octubre', 11: 'Noviembre', 12:'Diciembre'}


def datos_excel(dataframe : pd.DataFrame)->pd.DataFrame:
    """datos_excel
        Función encargada de crear las columnas a presentar en el reporte agregado
    Args:
        dataframe: El dataframe con la ejecución presupuestal agregada

    Returns:
        El dataframe con las columnas corregidas
    """
    # Realizamos los cambios en las columnas del dataframe
    dataframe.loc['TOTAL'] = [''] + dataframe.iloc[:,1:].sum().tolist()
    dataframe.loc['TOTAL','RUBRO'] = 'TOTAL DIP'
    dataframe.loc['TOTAL','DESCRIPCION'] = 'TOTAL DIP'
    
    # Calculamos el porcentaje de ejecución como un cociente y aplicamos un round para reducir el número de decimales que entrega la operación
    dataframe.loc['TOTAL','% COMPROMETIDO (B/A)'] = round((dataframe.loc['TOTAL','COMPROMISO (B)'] / dataframe.loc['TOTAL','APR. VIGENTE (A)']), 4)
    dataframe.loc['TOTAL','% EJECUCIÓN (C/A)'] = round((dataframe.loc['TOTAL','PAGOS (C)'] / dataframe.loc['TOTAL','APR. VIGENTE (A)']), 4)
    
    # Devolvemos el dataframe
    return dataframe

def ctr_reservas(dataframe : pd.DataFrame)->pd.DataFrame:
    """ctr_reservas

    Args:
        dataframe dataframe con la información de las reservas de la vigencia

    Returns:
        DataFrame con los cambios resultantes
    """
    
    # Filtramos el archivo de reservas solamente para la dirección de inclusión producitiva
    reserva_DIP = dataframe[dataframe['Dirección'] == 'Inclusión Productiva'].reset_index(drop=True)
    
    # Encontramos el valor con el total de las variables
    total_reserva = reserva_DIP['Valor Actual'].sum()
    
    # Seleccionamos unicamente las variables que necesitamos y las renombramos
    reserva_DIP = reserva_DIP[['Descripción Rubro','Numero Documento Soporte','Nombre Razon Social','Valor Actual','Valor Obligaciones','Valor Ordenes de Pago']]
    reserva_DIP = reserva_DIP.rename(columns={'Descripción Rubro':'PROGRAMA',
                                            'Numero Documento Soporte':'CONVENIO / CONTRATO',
                                            'Nombre Razon Social': 'TERCERO',
                                            'Valor Actual':'VALOR CONSTITUCION DE RESERVA',
                                            'Valor Obligaciones':'OBLIGADO',
                                            'Valor Ordenes de Pago':'PAGADO'})
    
    # Definimos un diccionario vacio donde vamos a guardar la información resultante de cada programa
    data = {}
    
    # Definimos un dataframe vacío donde se va a almacenar el dataframe resultante de todas las operaciones
    result = pd.DataFrame()
    
    # Vamos a iterar sobre cada uno de los programas para poder realizar los cálculos de forma desagregada
    for i in set(reserva_DIP['PROGRAMA'].values):
        # Guardamos en el diccionario el dataframe con la información de ese programa
        data[i] = reserva_DIP[reserva_DIP['PROGRAMA']==i]
        
        # Agrupamos el DataFrame de reservas por cada uno de los convenios/contratos que contiene
        data[i] = data[i].groupby('CONVENIO / CONTRATO', as_index=False).agg({
            'PROGRAMA':'first',
            'TERCERO':'first',
            'VALOR CONSTITUCION DE RESERVA':'sum',
            'OBLIGADO':'sum',
            'PAGADO':'sum'})
        
        # Eliminamos todos aquellos valores en el que la reserva sea igual a cero
        data[i] = data[i][data[i]['VALOR CONSTITUCION DE RESERVA']!=0].reset_index(drop = True)
        
        # Cálculamos los porcentajes de ejecución 
        data[i]['% DE RESERVA'] = data[i].apply(lambda x: round(int(x['VALOR CONSTITUCION DE RESERVA'])/ int(total_reserva) ,2), axis = 1) 
        data[i]['% EJECUCION'] = data[i].apply(lambda x: round(int(x['PAGADO']) / int(x['VALOR CONSTITUCION DE RESERVA']), 2), axis = 1)
        
        # Seleccionamos las columnas que necesitamos del dataframe una vez hemos realizado los calculos necesarios
        data[i] = data[i][['PROGRAMA','CONVENIO / CONTRATO','TERCERO','VALOR CONSTITUCION DE RESERVA','% DE RESERVA','OBLIGADO','PAGADO','% EJECUCION']]
        
        # Definimos el valor total de la reserva para el programa
        data[i].loc['TOTAL'] = [''] + data[i].iloc[:,1:].sum().tolist()
        
        # Realizamos cambios en los valores de columnas que se van a presentar en el informe sobre la ultima fila
        data[i].loc['TOTAL','PROGRAMA'] = f'Total {i}'
        data[i].loc['TOTAL','CONVENIO / CONTRATO'] = ''
        data[i].loc['TOTAL','TERCERO'] = ''
        
        # Añadimos los valores que van a aparecen en la ultima fila del dataframe con los totales consolidados
        data[i].loc['TOTAL','% DE RESERVA'] = round(int(data[i].loc['TOTAL','VALOR CONSTITUCION DE RESERVA']) / int(total_reserva), 4)
        data[i].loc['TOTAL','% EJECUCION'] = round(int(data[i].loc['TOTAL','PAGADO']) / int(data[i].loc['TOTAL','VALOR CONSTITUCION DE RESERVA']), 4)
        
        # Concatenamos en el DataFrame vacío toda la información que calculamos sobre cada programa
        result = pd.concat([result,data[i]])       

    # Al igual que cada programa añadimos los valores que van a aparecer en la ultima fila como un total desagregado por cada columna
    result.loc['TOTAL RESERVAS','PROGRAMA'] = 'TOTAL RESERVAS DIP'
    result.loc['TOTAL RESERVAS','VALOR CONSTITUCION DE RESERVA'] = reserva_DIP['VALOR CONSTITUCION DE RESERVA'].sum()
    result.loc['TOTAL RESERVAS','% DE RESERVA'] = 1
    result.loc['TOTAL RESERVAS','OBLIGADO'] = reserva_DIP['OBLIGADO'].sum()
    result.loc['TOTAL RESERVAS','PAGADO'] = reserva_DIP['PAGADO'].sum()
    
    # Calculamos el porcentaje de ejecución global de la dirección
    result.loc['TOTAL RESERVAS','% EJECUCION'] = round(int(result.loc['TOTAL RESERVAS','PAGADO']) / int(result.loc['TOTAL RESERVAS','VALOR CONSTITUCION DE RESERVA']), 2)
    
    # Devolvamos el DataFrame resultante
    return result

def export_report(archivo_de_datos: dict, path: str, month: int, year: int)-> None:
    """export_report
        función para exportar el reporte financiero a excel
    Args:
        archivo_de_datos (dict): diccionario con todos los dataframes modificados
        path (str): ruta donde se va a guardar el archivo
        month (int): número del mes actual
        year (int): número del año actual
    """
    # Estructuramos la ruta de la carpeta deon
    path_file = os.path.join(path,'Ejec Pptal '+ month_dict[month] + ' 2023.xlsx')
    # Creamos un archivo excel en el que se va a guardar la información
    writer = pd.ExcelWriter(path_file, engine='openpyxl')

    # Exportamos el dataframe con el reporte agregado a una hoja de excel
    archivo_de_datos['RG_Excel'].to_excel(writer, sheet_name='Ejecución Pptal Agregada DIP', index=False, startrow=3)

    # Obtener la hoja de Excel
    workbook = writer.book
    hoja = writer.sheets['Ejecución Pptal Agregada DIP']

    # Agregar los valores a las celdas referentes a la información del informe
    hoja['A1'] = 'Año fiscal:'
    hoja['B1'] = year
    hoja['A2'] = 'Vigencia:'
    hoja['B2'] = 'Actual'
    hoja['A3'] = 'Periodo:'
    hoja['B3'] = month_dict[month]

    # Guardamos el DataFrame con la información de reservas a una hoja del libro de excel
    archivo_de_datos['Reservas_Excel'].to_excel(writer, sheet_name='Control RVAS', index=False, startrow=1)
    
    # Accedemos a la hoja de reservas
    hoja2 = writer.sheets['Control RVAS']

    # Definimos el formato de color que tendran las columnas que vamos a modificar
    light_blue_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        
    # Añadimos el patrón de color a las columnas que contengan el valor total
    for row in hoja2.iter_rows():
        if 'Total' in str(row[0].value):
            for cell in row:
                cell.fill = light_blue_fill
                
    # Guardar el archivo
    writer.close()
    

